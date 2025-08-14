"""
Market Export Utils
市场分析结果导出工具 - 支持Excel、PDF、HTML、CSV等格式导出
"""

import os
import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime
import tempfile
import zipfile
import base64
from io import BytesIO
import streamlit as st

# 导入日志模块
from tradingagents.utils.logging_manager import get_logger
logger = get_logger('market_export_utils')

# 可选依赖导入
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl未安装，Excel导出功能不可用")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab未安装，PDF导出功能不可用")

try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.offline import plot
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    logger.warning("plotly未安装，图表导出功能受限")


class MarketExportManager:
    """市场分析结果导出管理器"""
    
    def __init__(self, export_dir: str = "./data/exports"):
        """
        初始化导出管理器
        
        Args:
            export_dir: 导出文件存储目录
        """
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        # 支持的导出格式
        self.supported_formats = {
            'excel': {'extension': 'xlsx', 'available': EXCEL_AVAILABLE},
            'csv': {'extension': 'csv', 'available': True},
            'json': {'extension': 'json', 'available': True},
            'html': {'extension': 'html', 'available': True},
            'pdf': {'extension': 'pdf', 'available': PDF_AVAILABLE}
        }
        
        logger.info(f"初始化市场导出管理器，导出目录: {export_dir}")
    
    def export_scan_results(self, scan_id: str, results_data: Dict[str, Any], 
                           format_type: str = 'excel', 
                           options: Optional[Dict] = None) -> Optional[str]:
        """
        导出分析结果
        
        Args:
            scan_id: 扫描ID
            results_data: 结果数据
            format_type: 导出格式 (excel, csv, json, html, pdf)
            options: 导出选项
            
        Returns:
            Optional[str]: 导出文件路径，失败返回None
        """
        
        try:
            # 检查格式支持
            if format_type not in self.supported_formats:
                raise ValueError(f"不支持的导出格式: {format_type}")
            
            format_info = self.supported_formats[format_type]
            if not format_info['available']:
                raise RuntimeError(f"{format_type}导出功能不可用，缺少必要依赖")
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"market_scan_{scan_id}_{timestamp}.{format_info['extension']}"
            export_path = self.export_dir / filename
            
            # 根据格式调用相应的导出方法
            if format_type == 'excel':
                success = self._export_to_excel(export_path, results_data, options)
            elif format_type == 'csv':
                success = self._export_to_csv(export_path, results_data, options)
            elif format_type == 'json':
                success = self._export_to_json(export_path, results_data, options)
            elif format_type == 'html':
                success = self._export_to_html(export_path, results_data, options)
            elif format_type == 'pdf':
                success = self._export_to_pdf(export_path, results_data, options)
            else:
                raise ValueError(f"未实现的导出格式: {format_type}")
            
            if success:
                logger.info(f"导出成功: {export_path}")
                return str(export_path)
            else:
                logger.error(f"导出失败: {format_type}")
                return None
                
        except Exception as e:
            logger.error(f"导出分析结果失败: {e}")
            return None
    
    def export_rankings_only(self, scan_id: str, rankings_data: List[Dict], 
                            format_type: str = 'csv') -> Optional[str]:
        """
        仅导出股票排名数据
        
        Args:
            scan_id: 扫描ID
            rankings_data: 排名数据
            format_type: 导出格式
            
        Returns:
            Optional[str]: 导出文件路径
        """
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            format_info = self.supported_formats.get(format_type, {'extension': 'csv'})
            filename = f"rankings_{scan_id}_{timestamp}.{format_info['extension']}"
            export_path = self.export_dir / filename
            
            # 转换为DataFrame
            df = pd.DataFrame(rankings_data)
            
            if format_type == 'csv':
                df.to_csv(export_path, index=False, encoding='utf-8-sig')
            elif format_type == 'excel' and EXCEL_AVAILABLE:
                df.to_excel(export_path, index=False)
            elif format_type == 'json':
                with open(export_path, 'w', encoding='utf-8') as f:
                    json.dump(rankings_data, f, ensure_ascii=False, indent=2)
            else:
                raise ValueError(f"不支持的格式: {format_type}")
            
            logger.info(f"排名数据导出成功: {export_path}")
            return str(export_path)
            
        except Exception as e:
            logger.error(f"导出排名数据失败: {e}")
            return None
    
    def create_export_package(self, scan_id: str, results_data: Dict[str, Any]) -> Optional[str]:
        """
        创建包含多种格式的导出包
        
        Args:
            scan_id: 扫描ID
            results_data: 结果数据
            
        Returns:
            Optional[str]: 压缩包路径
        """
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            package_name = f"market_analysis_package_{scan_id}_{timestamp}.zip"
            package_path = self.export_dir / package_name
            
            with zipfile.ZipFile(package_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                
                # 导出各种格式
                formats_to_export = ['json', 'csv']
                if EXCEL_AVAILABLE:
                    formats_to_export.append('excel')
                if PDF_AVAILABLE:
                    formats_to_export.append('pdf')
                
                for format_type in formats_to_export:
                    try:
                        export_path = self.export_scan_results(scan_id, results_data, format_type)
                        if export_path and os.path.exists(export_path):
                            # 添加到压缩包
                            arcname = os.path.basename(export_path)
                            zipf.write(export_path, arcname)
                            # 删除临时文件
                            os.remove(export_path)
                    except Exception as e:
                        logger.warning(f"导出{format_type}格式失败: {e}")
                        continue
                
                # 添加导出说明文件
                readme_content = self._generate_export_readme(scan_id, results_data)
                zipf.writestr("README.txt", readme_content)
            
            logger.info(f"导出包创建成功: {package_path}")
            return str(package_path)
            
        except Exception as e:
            logger.error(f"创建导出包失败: {e}")
            return None
    
    def _export_to_excel(self, file_path: Path, results_data: Dict, options: Optional[Dict]) -> bool:
        """导出为Excel格式"""
        
        if not EXCEL_AVAILABLE:
            return False
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                
                # 股票排名工作表
                if 'rankings' in results_data:
                    rankings_df = pd.DataFrame(results_data['rankings'])
                    if not rankings_df.empty:
                        # 重命名列
                        column_rename = {
                            'symbol': '股票代码',
                            'name': '股票名称',
                            'total_score': '综合评分',
                            'technical_score': '技术评分',
                            'fundamental_score': '基本面评分',
                            'current_price': '当前价格',
                            'change_percent': '涨跌幅',
                            'recommendation': '投资建议',
                            'target_price': '目标价格'
                        }
                        rankings_df = rankings_df.rename(columns=column_rename)
                        rankings_df.to_excel(writer, sheet_name='股票排名', index=False)
                        
                        # 格式化工作表
                        self._format_excel_sheet(writer.sheets['股票排名'], rankings_df)
                
                # 板块分析工作表
                if 'sectors' in results_data:
                    sectors_data = results_data['sectors']
                    if sectors_data:
                        sectors_list = []
                        for sector_name, sector_info in sectors_data.items():
                            sectors_list.append({
                                '板块名称': sector_name,
                                '涨跌幅(%)': sector_info.get('change_percent', 0),
                                '成交额(亿)': sector_info.get('volume', 0),
                                '活跃度': sector_info.get('activity_score', 0),
                                '推荐度': sector_info.get('recommendation_score', 0),
                                '龙头股票': sector_info.get('leading_stock', ''),
                                '推荐股票数': sector_info.get('recommended_count', 0)
                            })
                        
                        sectors_df = pd.DataFrame(sectors_list)
                        sectors_df.to_excel(writer, sheet_name='板块分析', index=False)
                        self._format_excel_sheet(writer.sheets['板块分析'], sectors_df)
                
                # 市场指标工作表
                if 'breadth' in results_data:
                    breadth_data = results_data['breadth']
                    breadth_list = []
                    for key, value in breadth_data.items():
                        if isinstance(value, (int, float)):
                            breadth_list.append({
                                '指标名称': self._translate_breadth_key(key),
                                '数值': value,
                                '单位': self._get_breadth_unit(key)
                            })
                    
                    if breadth_list:
                        breadth_df = pd.DataFrame(breadth_list)
                        breadth_df.to_excel(writer, sheet_name='市场指标', index=False)
                        self._format_excel_sheet(writer.sheets['市场指标'], breadth_df)
                
                # 执行摘要工作表
                if 'summary' in results_data:
                    summary_data = results_data['summary']
                    summary_content = self._format_summary_for_excel(summary_data)
                    
                    # 创建摘要DataFrame
                    summary_df = pd.DataFrame(summary_content)
                    summary_df.to_excel(writer, sheet_name='执行摘要', index=False)
                    self._format_excel_sheet(writer.sheets['执行摘要'], summary_df)
            
            return True
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            return False
    
    def _export_to_csv(self, file_path: Path, results_data: Dict, options: Optional[Dict]) -> bool:
        """导出为CSV格式"""
        
        try:
            # 主要导出股票排名数据
            if 'rankings' in results_data:
                rankings_df = pd.DataFrame(results_data['rankings'])
                if not rankings_df.empty:
                    # 重命名列为中文
                    column_rename = {
                        'symbol': '股票代码',
                        'name': '股票名称', 
                        'total_score': '综合评分',
                        'technical_score': '技术评分',
                        'fundamental_score': '基本面评分',
                        'current_price': '当前价格',
                        'change_percent': '涨跌幅',
                        'recommendation': '投资建议',
                        'target_price': '目标价格'
                    }
                    rankings_df = rankings_df.rename(columns=column_rename)
                    rankings_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"CSV导出失败: {e}")
            return False
    
    def _export_to_json(self, file_path: Path, results_data: Dict, options: Optional[Dict]) -> bool:
        """导出为JSON格式"""
        
        try:
            # 添加导出元数据
            export_data = {
                'export_info': {
                    'export_time': datetime.now().isoformat(),
                    'export_version': '1.0',
                    'source': 'TradingAgents-CN Market Analysis'
                },
                'scan_results': results_data
            }
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            return True
            
        except Exception as e:
            logger.error(f"JSON导出失败: {e}")
            return False
    
    def _export_to_html(self, file_path: Path, results_data: Dict, options: Optional[Dict]) -> bool:
        """导出为HTML格式"""
        
        try:
            html_content = self._generate_html_report(results_data)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return True
            
        except Exception as e:
            logger.error(f"HTML导出失败: {e}")
            return False
    
    def _export_to_pdf(self, file_path: Path, results_data: Dict, options: Optional[Dict]) -> bool:
        """导出为PDF格式"""
        
        if not PDF_AVAILABLE:
            return False
        
        try:
            doc = SimpleDocTemplate(str(file_path), pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # 添加标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.darkblue,
                spaceAfter=20,
                alignment=1  # 居中
            )
            
            story.append(Paragraph("全球市场分析报告", title_style))
            story.append(Spacer(1, 12))
            
            # 添加生成时间
            story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # 概览信息
            if 'total_stocks' in results_data:
                overview_data = [
                    ['指标', '数值'],
                    ['分析股票数', str(results_data.get('total_stocks', 0))],
                    ['推荐股票数', str(results_data.get('recommended_stocks', 0))],
                    ['实际成本', f"¥{results_data.get('actual_cost', 0):.2f}"],
                    ['扫描时长', str(results_data.get('scan_duration', '未知'))]
                ]
                
                overview_table = Table(overview_data)
                overview_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(Paragraph("扫描概览", styles['Heading2']))
                story.append(overview_table)
                story.append(Spacer(1, 20))
            
            # 股票排名表格
            if 'rankings' in results_data and results_data['rankings']:
                story.append(Paragraph("股票排名 (前20名)", styles['Heading2']))
                
                rankings_data = results_data['rankings'][:20]  # 只显示前20名
                
                # 准备表格数据
                table_data = [['排名', '股票代码', '股票名称', '综合评分', '投资建议']]
                
                for i, stock in enumerate(rankings_data, 1):
                    table_data.append([
                        str(i),
                        stock.get('symbol', ''),
                        stock.get('name', ''),
                        f"{stock.get('total_score', 0):.1f}",
                        stock.get('recommendation', '')
                    ])
                
                rankings_table = Table(table_data)
                rankings_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8)
                ]))
                
                story.append(rankings_table)
                story.append(PageBreak())
            
            # 执行摘要
            if 'summary' in results_data:
                story.append(Paragraph("执行摘要", styles['Heading2']))
                summary = results_data['summary']
                
                # 核心观点
                if 'key_insights' in summary:
                    story.append(Paragraph("核心观点:", styles['Heading3']))
                    insights = summary['key_insights']
                    if isinstance(insights, list):
                        for insight in insights:
                            story.append(Paragraph(f"• {insight}", styles['Normal']))
                    else:
                        story.append(Paragraph(str(insights), styles['Normal']))
                    story.append(Spacer(1, 12))
                
                # 投资建议
                if 'investment_recommendations' in summary:
                    story.append(Paragraph("投资建议:", styles['Heading3']))
                    recs = summary['investment_recommendations']
                    
                    if 'buy' in recs and recs['buy']:
                        story.append(Paragraph("推荐买入:", styles['Heading4']))
                        for stock in recs['buy'][:5]:
                            story.append(Paragraph(f"• {stock.get('name', '')} ({stock.get('symbol', '')})", styles['Normal']))
                    
                    story.append(Spacer(1, 12))
                
                # 风险提示
                if 'risk_factors' in summary:
                    story.append(Paragraph("风险提示:", styles['Heading3']))
                    risks = summary['risk_factors']
                    if isinstance(risks, list):
                        for risk in risks:
                            story.append(Paragraph(f"⚠️ {risk}", styles['Normal']))
                    else:
                        story.append(Paragraph(f"⚠️ {risks}", styles['Normal']))
            
            # 构建PDF
            doc.build(story)
            return True
            
        except Exception as e:
            logger.error(f"PDF导出失败: {e}")
            return False
    
    def _format_excel_sheet(self, worksheet, dataframe):
        """格式化Excel工作表"""
        
        if not EXCEL_AVAILABLE:
            return
        
        try:
            # 设置标题行格式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
        
        except Exception as e:
            logger.warning(f"格式化Excel工作表失败: {e}")
    
    def _generate_html_report(self, results_data: Dict) -> str:
        """生成HTML报告"""
        
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>全球市场分析报告</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 20px;
            line-height: 1.6;
        }}
        .header {{
            text-align: center;
            color: #2c3e50;
            border-bottom: 2px solid #3498db;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .overview {{
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 30px;
        }}
        .overview h2 {{
            color: #2c3e50;
            margin-top: 0;
        }}
        .metric {{
            display: inline-block;
            margin: 10px 20px;
            padding: 10px;
            background-color: white;
            border-radius: 5px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #3498db;
            color: white;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .buy-recommendation {{
            background-color: #d4edda;
            color: #155724;
            padding: 2px 8px;
            border-radius: 4px;
        }}
        .hold-recommendation {{
            background-color: #fff3cd;
            color: #856404;
            padding: 2px 8px;
            border-radius: 4px;
        }}
        .summary-section {{
            margin: 30px 0;
            padding: 20px;
            border-left: 4px solid #3498db;
            background-color: #f8f9fa;
        }}
        .insight {{
            margin: 10px 0;
            padding: 10px;
            background-color: white;
            border-radius: 5px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🌍 全球市场分析报告</h1>
        <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
        """
        
        # 添加概览信息
        html_content += """
    <div class="overview">
        <h2>📊 扫描概览</h2>
        """
        
        if 'total_stocks' in results_data:
            html_content += f"""
        <div class="metric">
            <strong>分析股票数:</strong> {results_data.get('total_stocks', 0)}只
        </div>
        <div class="metric">
            <strong>推荐股票数:</strong> {results_data.get('recommended_stocks', 0)}只
        </div>
        <div class="metric">
            <strong>实际成本:</strong> ¥{results_data.get('actual_cost', 0):.2f}
        </div>
        <div class="metric">
            <strong>分析时长:</strong> {results_data.get('scan_duration', '未知')}
        </div>
            """
        
        html_content += "</div>"
        
        # 添加股票排名表格
        if 'rankings' in results_data and results_data['rankings']:
            html_content += """
    <h2>📋 股票排名 (前50名)</h2>
    <table>
        <thead>
            <tr>
                <th>排名</th>
                <th>股票代码</th>
                <th>股票名称</th>
                <th>综合评分</th>
                <th>技术评分</th>
                <th>基本面评分</th>
                <th>当前价格</th>
                <th>涨跌幅</th>
                <th>投资建议</th>
            </tr>
        </thead>
        <tbody>
            """
            
            for i, stock in enumerate(results_data['rankings'][:50], 1):
                recommendation = stock.get('recommendation', '')
                rec_class = 'buy-recommendation' if recommendation == '买入' else 'hold-recommendation'
                
                html_content += f"""
            <tr>
                <td>{i}</td>
                <td>{stock.get('symbol', '')}</td>
                <td>{stock.get('name', '')}</td>
                <td>{stock.get('total_score', 0):.1f}</td>
                <td>{stock.get('technical_score', 0):.1f}</td>
                <td>{stock.get('fundamental_score', 0):.1f}</td>
                <td>¥{stock.get('current_price', 0):.2f}</td>
                <td>{stock.get('change_percent', 0):+.2f}%</td>
                <td><span class="{rec_class}">{recommendation}</span></td>
            </tr>
                """
            
            html_content += """
        </tbody>
    </table>
            """
        
        # 添加执行摘要
        if 'summary' in results_data:
            summary = results_data['summary']
            html_content += """
    <div class="summary-section">
        <h2>📑 执行摘要</h2>
            """
            
            # 核心观点
            if 'key_insights' in summary:
                html_content += "<h3>💡 核心观点</h3>"
                insights = summary['key_insights']
                if isinstance(insights, list):
                    for insight in insights:
                        html_content += f'<div class="insight">• {insight}</div>'
                else:
                    html_content += f'<div class="insight">{insights}</div>'
            
            # 投资建议
            if 'investment_recommendations' in summary:
                html_content += "<h3>🎯 投资建议</h3>"
                recs = summary['investment_recommendations']
                
                if 'buy' in recs and recs['buy']:
                    html_content += "<h4>💚 推荐买入:</h4><ul>"
                    for stock in recs['buy'][:10]:
                        html_content += f"<li>{stock.get('name', '')} ({stock.get('symbol', '')})</li>"
                    html_content += "</ul>"
            
            html_content += "</div>"
        
        html_content += """
</body>
</html>
        """
        
        return html_content
    
    def _format_summary_for_excel(self, summary_data: Dict) -> List[Dict]:
        """格式化摘要数据用于Excel导出"""
        
        summary_content = []
        
        # 核心观点
        if 'key_insights' in summary_data:
            insights = summary_data['key_insights']
            if isinstance(insights, list):
                for i, insight in enumerate(insights, 1):
                    summary_content.append({
                        '类型': '核心观点',
                        '序号': i,
                        '内容': insight
                    })
            else:
                summary_content.append({
                    '类型': '核心观点',
                    '序号': 1,
                    '内容': str(insights)
                })
        
        # 投资建议
        if 'investment_recommendations' in summary_data:
            recs = summary_data['investment_recommendations']
            
            if 'buy' in recs and recs['buy']:
                for i, stock in enumerate(recs['buy'][:10], 1):
                    summary_content.append({
                        '类型': '推荐买入',
                        '序号': i,
                        '内容': f"{stock.get('name', '')} ({stock.get('symbol', '')})"
                    })
        
        # 风险提示
        if 'risk_factors' in summary_data:
            risks = summary_data['risk_factors']
            if isinstance(risks, list):
                for i, risk in enumerate(risks, 1):
                    summary_content.append({
                        '类型': '风险提示',
                        '序号': i,
                        '内容': risk
                    })
            else:
                summary_content.append({
                    '类型': '风险提示',
                    '序号': 1,
                    '内容': str(risks)
                })
        
        return summary_content
    
    def _translate_breadth_key(self, key: str) -> str:
        """翻译市场广度指标键名"""
        
        translation_map = {
            'up_ratio': '上涨股票占比',
            'activity_index': '成交活跃度',
            'net_inflow': '资金净流入',
            'sentiment_index': '市场情绪指数',
            'market_strength': '市场强度',
            'limit_up_count': '涨停股票数',
            'limit_down_count': '跌停股票数',
            'new_high_count': '新高股票数',
            'new_low_count': '新低股票数',
            'high_volume_count': '放量股票数'
        }
        
        return translation_map.get(key, key)
    
    def _get_breadth_unit(self, key: str) -> str:
        """获取市场广度指标单位"""
        
        unit_map = {
            'up_ratio': '%',
            'activity_index': '分',
            'net_inflow': '亿元',
            'sentiment_index': '分',
            'market_strength': '分',
            'limit_up_count': '只',
            'limit_down_count': '只',
            'new_high_count': '只',
            'new_low_count': '只',
            'high_volume_count': '只'
        }
        
        return unit_map.get(key, '')
    
    def _generate_export_readme(self, scan_id: str, results_data: Dict) -> str:
        """生成导出包说明文件"""
        
        readme_content = f"""
TradingAgents-CN 全球市场分析导出包
=====================================

扫描ID: {scan_id}
导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

文件说明:
--------
- market_scan_{scan_id}_*.xlsx: Excel格式的完整分析报告，包含多个工作表
- market_scan_{scan_id}_*.csv: CSV格式的股票排名数据，适合数据分析
- market_scan_{scan_id}_*.json: JSON格式的原始数据，包含所有分析结果
- market_scan_{scan_id}_*.html: HTML格式的可视化报告，可在浏览器中查看
- market_scan_{scan_id}_*.pdf: PDF格式的专业报告，适合打印和分享

数据概览:
--------
"""
        
        if 'total_stocks' in results_data:
            readme_content += f"""
分析股票数: {results_data.get('total_stocks', 0)}只
推荐股票数: {results_data.get('recommended_stocks', 0)}只
实际成本: ¥{results_data.get('actual_cost', 0):.2f}
分析时长: {results_data.get('scan_duration', '未知')}
"""
        
        readme_content += f"""

数据结构说明:
-----------
1. 股票排名 (rankings): 按综合评分排序的股票列表
2. 板块分析 (sectors): 各板块表现和推荐股票
3. 市场广度 (breadth): 市场整体健康度指标
4. 执行摘要 (summary): AI生成的分析总结和投资建议

注意事项:
--------
- 本分析结果仅供参考，不构成投资建议
- 投资有风险，决策需谨慎
- 数据来源于扫描时的市场状况，可能存在时效性

技术支持:
--------
TradingAgents-CN 项目
GitHub: https://github.com/your-repo/TradingAgents-CN
"""
        
        return readme_content
    
    def get_export_history(self, limit: int = 50) -> List[Dict[str, Any]]:
        """获取导出历史记录"""
        
        try:
            export_files = []
            
            for file_path in self.export_dir.iterdir():
                if file_path.is_file() and not file_path.name.startswith('.'):
                    stat = file_path.stat()
                    
                    export_files.append({
                        'filename': file_path.name,
                        'size': stat.st_size,
                        'created_time': datetime.fromtimestamp(stat.st_mtime),
                        'file_path': str(file_path)
                    })
            
            # 按创建时间排序
            export_files.sort(key=lambda x: x['created_time'], reverse=True)
            
            return export_files[:limit]
            
        except Exception as e:
            logger.error(f"获取导出历史失败: {e}")
            return []
    
    def cleanup_old_exports(self, days: int = 30) -> int:
        """清理旧的导出文件"""
        
        try:
            cutoff_time = datetime.now().timestamp() - (days * 24 * 3600)
            cleaned_count = 0
            
            for file_path in self.export_dir.iterdir():
                if file_path.is_file():
                    if file_path.stat().st_mtime < cutoff_time:
                        file_path.unlink()
                        cleaned_count += 1
            
            logger.info(f"清理了{cleaned_count}个旧导出文件")
            return cleaned_count
            
        except Exception as e:
            logger.error(f"清理旧导出文件失败: {e}")
            return 0


# Streamlit集成函数
def render_export_interface(scan_id: str, results_data: Dict[str, Any]):
    """渲染导出界面 - 增强版本"""
    
    # 导入增强版导出功能
    try:
        from utils.enhanced_market_export_utils import render_enhanced_export_interface
        render_enhanced_export_interface(scan_id, results_data)
        logger.info("使用增强版导出界面")
        
    except ImportError as e:
        logger.warning(f"增强版导出功能导入失败，使用标准版本: {e}")
        # 回退到标准版本
        render_standard_export_interface(scan_id, results_data)


def render_standard_export_interface(scan_id: str, results_data: Dict[str, Any]):
    """渲染Streamlit导出界面"""
    
    st.markdown("### 📤 结果导出")
    st.caption("将分析结果导出为不同格式，便于后续使用和分享")
    
    # 初始化导出管理器
    if 'export_manager' not in st.session_state:
        st.session_state.export_manager = MarketExportManager()
    
    export_manager = st.session_state.export_manager
    
    # 导出选项
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        if st.button("📊 Excel格式", use_container_width=True, 
                    disabled=not EXCEL_AVAILABLE,
                    help="导出为Excel文件，包含多个工作表" if EXCEL_AVAILABLE else "需要安装openpyxl"):
            with st.spinner("正在生成Excel文件..."):
                export_path = export_manager.export_scan_results(scan_id, results_data, 'excel')
                if export_path:
                    st.success("✅ Excel文件生成成功")
                    # 提供下载链接
                    with open(export_path, 'rb') as f:
                        st.download_button(
                            label="📥 下载Excel文件",
                            data=f.read(),
                            file_name=os.path.basename(export_path),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.error("❌ Excel导出失败")
    
    with col2:
        if st.button("📄 CSV格式", use_container_width=True,
                    help="导出股票排名数据为CSV文件"):
            with st.spinner("正在生成CSV文件..."):
                export_path = export_manager.export_scan_results(scan_id, results_data, 'csv')
                if export_path:
                    st.success("✅ CSV文件生成成功")
                    with open(export_path, 'rb') as f:
                        st.download_button(
                            label="📥 下载CSV文件",
                            data=f.read(),
                            file_name=os.path.basename(export_path),
                            mime="text/csv"
                        )
                else:
                    st.error("❌ CSV导出失败")
    
    with col3:
        if st.button("🌐 HTML报告", use_container_width=True,
                    help="生成可在浏览器中查看的HTML报告"):
            with st.spinner("正在生成HTML报告..."):
                export_path = export_manager.export_scan_results(scan_id, results_data, 'html')
                if export_path:
                    st.success("✅ HTML报告生成成功")
                    with open(export_path, 'rb') as f:
                        st.download_button(
                            label="📥 下载HTML报告",
                            data=f.read(),
                            file_name=os.path.basename(export_path),
                            mime="text/html"
                        )
                else:
                    st.error("❌ HTML导出失败")
    
    with col4:
        if st.button("📋 JSON数据", use_container_width=True,
                    help="导出原始JSON数据"):
            with st.spinner("正在生成JSON文件..."):
                export_path = export_manager.export_scan_results(scan_id, results_data, 'json')
                if export_path:
                    st.success("✅ JSON文件生成成功")
                    with open(export_path, 'rb') as f:
                        st.download_button(
                            label="📥 下载JSON文件",
                            data=f.read(),
                            file_name=os.path.basename(export_path),
                            mime="application/json"
                        )
                else:
                    st.error("❌ JSON导出失败")
    
    with col5:
        if st.button("📦 完整包", use_container_width=True,
                    help="创建包含所有格式的压缩包"):
            with st.spinner("正在创建导出包..."):
                package_path = export_manager.create_export_package(scan_id, results_data)
                if package_path:
                    st.success("✅ 导出包创建成功")
                    with open(package_path, 'rb') as f:
                        st.download_button(
                            label="📥 下载完整包",
                            data=f.read(),
                            file_name=os.path.basename(package_path),
                            mime="application/zip"
                        )
                else:
                    st.error("❌ 导出包创建失败")
    
    # PDF导出单独处理
    if PDF_AVAILABLE:
        col_pdf, col_empty = st.columns([1, 4])
        with col_pdf:
            if st.button("📄 PDF报告", use_container_width=True,
                        help="生成专业PDF报告"):
                with st.spinner("正在生成PDF报告..."):
                    export_path = export_manager.export_scan_results(scan_id, results_data, 'pdf')
                    if export_path:
                        st.success("✅ PDF报告生成成功")
                        with open(export_path, 'rb') as f:
                            st.download_button(
                                label="📥 下载PDF报告",
                                data=f.read(),
                                file_name=os.path.basename(export_path),
                                mime="application/pdf"
                            )
                    else:
                        st.error("❌ PDF导出失败")
    else:
        st.info("💡 安装reportlab库可启用PDF导出功能")
    
    # 显示功能说明
    with st.expander("📖 导出格式说明"):
        st.markdown("""
        **各种导出格式的特点:**
        
        - **📊 Excel**: 完整的多工作表报告，适合深度分析和数据处理
        - **📄 CSV**: 纯数据格式，适合导入其他分析工具
        - **🌐 HTML**: 可视化网页报告，可在浏览器中查看和分享
        - **📋 JSON**: 原始数据格式，适合程序化处理
        - **📄 PDF**: 专业打印报告，适合正式文档存档
        - **📦 完整包**: 包含所有格式的压缩包，一次下载全部文件
        
        **建议用途:**
        - 快速查看：HTML报告
        - 数据分析：Excel或CSV
        - 存档分享：PDF报告
        - 程序处理：JSON数据
        """)


# 示例用法
if __name__ == "__main__":
    # 测试导出功能
    export_manager = MarketExportManager()
    
    # 模拟结果数据
    test_results = {
        'total_stocks': 100,
        'recommended_stocks': 25,
        'actual_cost': 12.50,
        'scan_duration': '15分钟',
        'rankings': [
            {
                'symbol': '000001',
                'name': '测试股票1',
                'total_score': 85.5,
                'technical_score': 80.0,
                'fundamental_score': 90.0,
                'current_price': 25.50,
                'change_percent': 2.5,
                'recommendation': '买入',
                'target_price': 30.0
            }
        ],
        'sectors': {
            '科技': {
                'change_percent': 2.5,
                'volume': 500,
                'activity_score': 75,
                'recommendation_score': 80
            }
        },
        'breadth': {
            'up_ratio': 65.5,
            'sentiment_index': 70.2,
            'market_strength': 68.0
        },
        'summary': {
            'key_insights': ['市场表现良好', '科技板块领涨'],
            'investment_recommendations': {
                'buy': [{'name': '测试股票1', 'symbol': '000001'}]
            },
            'risk_factors': ['市场波动风险']
        }
    }
    
    # 测试各种导出格式
    scan_id = 'test_scan_001'
    
    for format_type in ['json', 'csv']:
        print(f"测试{format_type}导出...")
        result = export_manager.export_scan_results(scan_id, test_results, format_type)
        print(f"{format_type}导出结果: {result}")
    
    print("导出测试完成")